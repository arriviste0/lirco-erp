import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook


def write_header(ws, title):
    ws["A1"] = title
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def write_kv(ws, start_row, data):
    row = start_row
    for key, value in data:
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1
    return row


def extract_cost_sheet_summary(cost_sheet):
    data = cost_sheet.get("data") or {}
    job_setup = data.get("jobSetup") or {}
    summary = data.get("summary") or {}
    totals = summary.get("totals") or {}
    cost_summary = summary.get("costSummary") or {}
    return {
        "costSheetNo": cost_sheet.get("costSheetNo", ""),
        "sheetDate": cost_sheet.get("sheetDate", ""),
        "customerName": job_setup.get("customerName", ""),
        "product": job_setup.get("product", ""),
        "totalQty": totals.get("totalQty", ""),
        "sellingRatePerPiece": cost_summary.get("sellingRatePerPiece", ""),
        "sellingRatePerKg": cost_summary.get("sellingRatePerKg", ""),
    }


def main():
    try:
        payload = json.load(sys.stdin)
    except json.JSONDecodeError:
        print(json.dumps({"ok": False, "error": "Invalid JSON payload"}))
        return 1

    output_path = payload.get("outputPath")
    if not output_path:
        print(json.dumps({"ok": False, "error": "Missing outputPath"}))
        return 1

    order = payload.get("order") or {}
    quotation = payload.get("quotation") or {}
    cost_sheets = payload.get("costSheets") or []
    inquiries = payload.get("inquiries") or []

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Order"
    write_header(ws, "Work Order")

    row = 4
    row = write_kv(
        ws,
        row,
        [
            ("Offer No", order.get("offerNo", "")),
            ("Offer Date", order.get("offerDate", "")),
            ("PO Number", order.get("poNumber", "")),
            ("PO Date", order.get("poDate", "")),
            ("Party Name", order.get("partyName", "")),
            ("Cost Sheet Nos", ", ".join(order.get("costSheetNos", []) or [])),
        ],
    )

    row += 1
    row = write_kv(
        ws,
        row,
        [
            ("Quotation Ref No", quotation.get("refNo", "")),
            ("Quotation Ref Date", quotation.get("refDate", "")),
            ("Party Address", quotation.get("partyAddress", "")),
            ("Contact No", quotation.get("contactNo", "")),
            ("Kind Attn", quotation.get("kindAttn", "")),
        ],
    )

    if inquiries:
        row += 2
        ws[f"A{row}"] = "Inquiries"
        row += 1
        ws[f"A{row}"] = "SL No"
        ws[f"B{row}"] = "Item"
        ws[f"C{row}"] = "Width"
        ws[f"D{row}"] = "Thickness"
        ws[f"E{row}"] = "Length"
        ws[f"F{row}"] = "Quantity"
        row += 1
        for inquiry in inquiries:
            ws[f"A{row}"] = inquiry.get("slNo", "")
            ws[f"B{row}"] = inquiry.get("item", "")
            ws[f"C{row}"] = inquiry.get("widthOd", "")
            ws[f"D{row}"] = inquiry.get("thickness", "")
            ws[f"E{row}"] = inquiry.get("length", "")
            ws[f"F{row}"] = inquiry.get("quantity", "")
            row += 1

    if cost_sheets:
        summary_ws = wb.create_sheet(title="Cost Sheets")
        summary_ws["A1"] = "Cost Sheet Summary"
        summary_ws["A3"] = "Cost Sheet No"
        summary_ws["B3"] = "Date"
        summary_ws["C3"] = "Customer"
        summary_ws["D3"] = "Product"
        summary_ws["E3"] = "Total Qty"
        summary_ws["F3"] = "Sell Rate/Pc"
        summary_ws["G3"] = "Sell Rate/Kg"
        row = 4
        for cost_sheet in cost_sheets:
            summary = extract_cost_sheet_summary(cost_sheet)
            summary_ws[f"A{row}"] = summary["costSheetNo"]
            summary_ws[f"B{row}"] = summary["sheetDate"]
            summary_ws[f"C{row}"] = summary["customerName"]
            summary_ws[f"D{row}"] = summary["product"]
            summary_ws[f"E{row}"] = summary["totalQty"]
            summary_ws[f"F{row}"] = summary["sellingRatePerPiece"]
            summary_ws[f"G{row}"] = summary["sellingRatePerKg"]
            row += 1

    wb.save(output_path)
    print(json.dumps({"ok": True, "filePath": output_path}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
