import copy
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import load_workbook
import openpyxl
from win32com.client import constants
import pandas as pd
import json
from datetime import datetime
from collections import defaultdict
import shutil
import os
import xlwings as xw
from num2words import num2words
import datetime as dt
import builtins

def safe_input(prompt="", default=""):
    try:
        return builtins.input(prompt)
    except EOFError:
        print(" No input available. Using default.")
        if default != "":
            return default
        # If the prompt looks numeric, return a safe numeric fallback.
        return "1" if any(ch.isdigit() for ch in str(prompt)) else ""

# Route all input calls through safe_input for non-interactive runs.
input = safe_input

def choose_from_list(prompt, options):
    print(prompt)
    for i, option in enumerate(options, start=1):
        print(f"{i}. {option}")
    while True:
        try:
            choice = input("Enter choice number: ").strip()
        except EOFError:
            if not options:
                raise
            print(" No input available. Defaulting to option 1.")
            return options[0]
        if choice == "":
            if not options:
                raise ValueError("No options provided.")
            print(" No choice provided. Defaulting to option 1.")
            return options[0]
        if choice.isdigit():
            index = int(choice) - 1
            if 0 <= index < len(options):
                return options[index]
        print(" Invalid choice. Please enter a valid number.")
# def choose_from_list(prompt, options):
#     print(prompt)
#     for i, option in enumerate(options, start=1):
#         print(f"{i}. {option}")
#     while True:
#         choice = input("Enter choice number: ").strip()
#         if choice.isdigit():
#             index = int(choice) - 1
#             if 0 <= index < len(options):
#                 return options[index]
#         print(" Invalid choice. Please enter a valid number.")


        
def get_valid_int(prompt, min_val=None, max_val=None):
    while True:
        try:
            raw = input(prompt).strip()
            if raw == "":
                return min_val if min_val is not None else 0
            val = int(raw)
            if min_val is not None and val < min_val:
                print(f"Enter a value ≥ {min_val}")
            elif max_val is not None and val > max_val:
                print(f"Enter a value ≤ {max_val}")
            else:
                return val
        except ValueError:
            print("Invalid input. Please enter an integer.")

def get_valid_float(prompt, min_val=None, max_val=None):
    while True:
        try:
            raw = input(prompt).strip()
            if raw == "":
                return float(min_val) if min_val is not None else 0.0
            val = float(raw)
            if min_val is not None and val < min_val:
                print(f"Enter a value ≥ {min_val}")
            elif max_val is not None and val > max_val:
                print(f"Enter a value ≤ {max_val}")
            else:
                return val
        except ValueError:
            print("Invalid input. Please enter a number (decimal allowed).")

def get_valid_date(prompt):
    while True:
        date_str = input(prompt).strip()
        if date_str == "":
            return datetime.today().strftime("%d-%b-%Y")
        try:
            date_obj = dt.datetime.strptime(date_str, "%d.%m.%y")
            return date_obj.strftime("%d-%b-%Y")  # Example: 19/Jul/2025
        except ValueError:
            print("Invalid format. Please enter date as dd-mm-yy (e.g., 19-07-25)")

def packing_details_def (product,profile_length,num_profiles,angle_weld_type,design,description,profile_width, nail_length):
    global grand_total_channel_kg
    global grand_total_angle_kg
    global final_summary
    def calculate_crate_width(arrangement, profile_thickness, flange_height, num_profiles,profile_width):
        current_width = profile_thickness
        if arrangement == 'single':
            for i in range(1, num_profiles):
                current_width += flange_height
        elif arrangement == 'double':
            for i in range(2, num_profiles + 1):
                if i % 2 == 1:
                    current_width += profile_thickness
                else:
                    current_width += flange_height
        current_width =  c_channel_height * 2  + current_width        
        return current_width
    def auto_fill_crates(total, max_per_crate, max_width):
        crates = []
        max_fit = 1
        while max_fit < max_per_crate:
            if calculate_crate_width(arrangement, profile_thickness, flange_height, max_fit + 1,profile_width) > max_width:
                break
            max_fit += 1
        while total > 0:
            count = min(max_fit, total)
            crates.append(count)
            total -= count
        return crates
    def calculate_crate_dimensions(per_crate_profiles, flange_height, profile_thickness):
        # Try different number of rows until we satisfy the crate height = 90% of crate width constraint
        for rows in range(1, per_crate_profiles + 1):
            cols = math.ceil(per_crate_profiles / rows)
            
            # Calculate width and height based on pipe dimensions
            crate_width = cols * flange_height
            crate_height = rows * profile_thickness

            # Check the 90% height condition
            if math.isclose(crate_height, 0.9 * crate_width, rel_tol=0.05):  # ±5% tolerance
                return {
                    "rows": rows,
                    "columns": cols,
                    "crate_width": round(crate_width, 2),
                    "crate_height": round(crate_height, 2)     }

        # If no such condition met, return the closest arrangement (brute force fallback)
        rows = math.ceil(math.sqrt(per_crate_profiles))
        cols = math.ceil(per_crate_profiles / rows)
        crate_width = cols * flange_height
        crate_height = rows * profile_thickness
        return {
            "rows": rows,
            "columns": cols,
            "crate_width": round(crate_width, 2),
            "crate_height": round(crate_height, 2),
            "note": "Condition not exactly met. Best approximate layout provided."    } 
    # -------- Design Configuration Dictionary -------- #
    define_packing = {
        'CE_pack': {
            'elex_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'single','flange_height': 16,'ce_height': 60, 'channel_height': 40, 'channel_width': 75, 'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 5},
            'elexcut_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'single','flange_height': 50,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 5}, 
            'elex_nor_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'single','flange_height': 16,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 5}, 
            'bhel_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'double','flange_height': 3,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 77}, 
            'bhel387_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'double','flange_height': 5,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 20}, 
            'thermax20_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'double','flange_height': 3,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 20},
            'sitson_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'single','flange_height': 16,'ce_height': 60, 'channel_height': 40, 'channel_width': 75, 'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 5},
            'thermax30_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 500,'max_crate_width': 1200,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'double','flange_height': 3,'ce_height': 50,'channel_height': 40,'channel_width': 75,'max_ce_crate': 42,
                'crate_height_space' : 20, 'extra_space_packing': 30},   },
        'DE_pack': {
            'std_pack': {
                'max_u_gap': 3200,'ce_ext_fr_u': 700,'max_crate_width': 1150,'channel_wt_per_m': 4.1,'angle_wt_per_m': 2.6,
                'arrange_ce': 'single','flange_height': 5,'ce_height': 65, 'channel_height': 40, 'channel_width': 75, 'max_ce_crate': 500,
                'crate_height_space' : 20, 'extra_space_packing': 5}, }    }
    
    # -------- Extract Parameters -------- #
    
    
    if product == "CE":
        pack = "CE_pack"
    else:pack = "DE_pack"   
    params = define_packing[pack][design]
    c_channel_height = params['channel_height']
    c_channel_width = params['channel_width']
    if product == "CE":
        flange_height = params['flange_height']
        profile_thickness = params['ce_height'] # to measure width of crate
        profile_height = profile_width + params['extra_space_packing'] # to measure crate height
        max_crate_width = params['max_crate_width']
        max_profiles_per_crate = params['max_ce_crate']
    else:
        part = round((nail_length / 2),0) - 8
        strip_width = get_valid_float(f"Partition strips width in crate for of DE: ({part} to 90): ", part, 90)
        flange_height =  profile_width + 0.5 + strip_width # Total Space required to arrange pipes side by side 
        max_crate_width = get_valid_int("Input maximum crate width for DE: (200 to 1300): ", 200, 1300)
        inner_width = max_crate_width - 2 * c_channel_height
        max_crate_height = inner_width * 0.90
        profile_thickness = profile_width + 0.5  # to measure width of crate
        horizontal_de_no = round(inner_width / flange_height,0)
        vertical_de_no = round(max_crate_height/profile_thickness,0)
        max_profiles_per_crate = round(horizontal_de_no * vertical_de_no,0)
        total_de_nos = num_profiles 
        if total_de_nos > max_profiles_per_crate:
            crates_nos = math.ceil(total_de_nos / max_profiles_per_crate)
            per_crate_pipes = math.ceil(total_de_nos / crates_nos)
            result = calculate_crate_dimensions(per_crate_pipes, flange_height, profile_thickness)
        else:
           crates_nos = 1
           per_crate_pipes = total_de_nos
           result = calculate_crate_dimensions(per_crate_pipes, flange_height, profile_thickness)
        profile_height = result['crate_height'] 
            
    max_u_gap = params['max_u_gap']
    ce_ext = params['ce_ext_fr_u']
    channel_wt_per_m = params['channel_wt_per_m']
    angle_wt_per_m = params['angle_wt_per_m']
    arrangement = params['arrange_ce']
    crate_height_space = params['crate_height_space'] 
    print(f"Prevoius U Gaps {previous_u_gap} Current Length : {profile_length}")
       
    crate_internal_length = profile_length - 2 * ce_ext
    num_squares = math.ceil(crate_internal_length / max_u_gap) + 1
    if num_squares > 1:
        gap_in_u = math.ceil((crate_internal_length / (num_squares - 1)) / 10) * 10
    else:
        gap_in_u = 0
    adjusted_ce_ext = (profile_length - (num_squares-1) * gap_in_u) / 2
    
    # -------- Crate Geometry -------- #
    while True:
        # user_input = input("Enter ce_extended from U value (or type 'pass' to exit): ").strip()
        user_input = input(f"Enter ce_extended value [default: {ce_ext}] or type 'e' to exit: ").strip()
        
        if user_input.lower() == 'e' or user_input.lower() == "" :
            print("Exiting input loop.")
            break
        try:
            ce_ext = float(user_input)  # take ce_ext from user
            crate_internal_length = profile_length - 2 * ce_ext
            num_squares = math.ceil(crate_internal_length / max_u_gap) + 1
            if num_squares > 1:
                gap_in_u = math.ceil((crate_internal_length / (num_squares - 1)) / 10) * 10
            else:
                gap_in_u = 0
            adjusted_ce_ext = (profile_length - (num_squares-1) * gap_in_u) / 2
            
            # Print output
            print(f"Results for ce_ext input {ce_ext} mm:")
            print(f"  Number of Squares (U frames): {num_squares}")
            print(f" Gap between Squares: {gap_in_u} mm")
            print(f" Recalculated Effective ce_ext: {round(adjusted_ce_ext, 2)} mm")
            print("-" * 50)
        except ValueError:
            print("Invalid input. Please enter a number or 'pass'.")
    
    # -------- Profiles per Crate Calculation & Crate dimension -------- #
    
    crate_length = profile_length - adjusted_ce_ext * 2
    if num_squares > 1:
        actual_gap_in_u = round(
            ((crate_length - (num_squares * c_channel_width)) / (num_squares - 1)), 0
        )
    else:
        actual_gap_in_u = 0
    crate_height = profile_height + c_channel_height * 2 + crate_height_space
    
    previous_u_gap.append(gap_in_u)
    
    # -- INITIAL FULL AUTO DISTRIBUTION --
    if product == "CE":
       crate_inputs = auto_fill_crates(num_profiles, max_profiles_per_crate, max_crate_width)
       while True:
           # BUILD SUMMARY
           crate_summary = {}
           crate_type_counter = 1
           for profiles in crate_inputs:
               width = calculate_crate_width(arrangement, profile_thickness, flange_height, profiles,profile_width)
               key = (profiles, round(width, 2))
               if key not in crate_summary:
                   crate_summary[key] = {'type_id': f'Crate Type {crate_type_counter}', 'count': 1}
                   crate_type_counter += 1
               else:
                   crate_summary[key]['count'] += 1
           final_summary = {}
           for (profiles, width), info in crate_summary.items():
               final_summary[info['type_id']] = {
                   'no_of_crates': info['count'],
                   'profiles_per_crate': profiles,
                   'crate_width': width,
                   'crate_length': crate_length, 
                   'crate_height' : crate_height, 
                   'inner_crate_height': profile_height,
                   'no_of_U_in_a_crate':num_squares,
                   'angle_weld_diagonal_or_box' : angle_weld_type,
                   'gap_between_u':actual_gap_in_u,
                   'channel_wt_per_m': channel_wt_per_m,
                   'angle_wt_per_m' : angle_wt_per_m,
                   'length_of_profiles': profile_length,
                   'flange_height': flange_height,
                   'width_of_profile': profile_width,
                   'description': description }
       
           # DISPLAY SUMMARY
           print("\n Crate Type Summary:")
           print(json.dumps(final_summary, indent=4))
       
           # Ask if user wants to modify
           repeat = input("\n Would you like to modify crate types again? (y / n): ").strip().lower()
           if repeat == 'n'or repeat == "" :
               print("\n Crate planning finalized.")
               break
       
           # --- MODIFICATION SECTION (if 'yes') ---
           print("\n Enter custom crate types as 'profiles crates'. Type 'pass' to auto-calculate the rest.\n")
           custom_inputs = []
           remaining_profiles = num_profiles
       
           while True:
               user_input = input("Enter crate type, pcs in a crate, no of crates (e.g., '40 34') or 'pass': ").strip().lower()
               if user_input == 'pass':
                   break
               try:
                   parts = user_input.split()
                   if len(parts) != 2:
                       print("Format should be: profiles crates")
                       continue
                   profiles = int(parts[0])
                   count = int(parts[1])
                   # if profiles > max_profiles_per_crate:
                   #     print(f"Cannot exceed {max_profiles_per_crate} profiles per crate.") gross_weight_per_crate
                   #     continue
                   width_check = calculate_crate_width(arrangement, profile_thickness, flange_height, profiles,profile_width)
                   if width_check > max_crate_width:
                       print(f"Crate width {width_check:.2f} mm exceeds limit of {max_crate_width} mm.")
                       continue
                   total_needed = profiles * count
                   if total_needed > remaining_profiles:
                       print(f"Only {remaining_profiles} profiles remaining.")
                       continue
                   custom_inputs.extend([profiles] * count)
                   remaining_profiles -= total_needed
               except ValueError:
                   print(" Invalid input. Please enter numbers or 'pass'.")
           if remaining_profiles > 0:
               print(f"\n Automatically placing remaining {remaining_profiles} profiles...")
               custom_inputs.extend(auto_fill_crates(remaining_profiles, max_profiles_per_crate, max_crate_width))
           # Replace main crate_inputs with the modified list
           crate_inputs = custom_inputs
           
           
    else:   # If Discharge Electrode
       final_summary = {}
       width = result['crate_width'] + c_channel_height * 2
       crate_height = result['crate_height'] + c_channel_height * 2
       final_summary['Crate Type 1'] = {
               'no_of_crates': crates_nos,
               'profiles_per_crate': per_crate_pipes,
               'crate_width': width,
               'crate_length': crate_length, 
               'crate_height' : crate_height, 
               'inner_crate_height': result['crate_height'],
               'no_of_U_in_a_crate':num_squares,
               'angle_weld_diagonal_or_box' : angle_weld_type,
               'gap_between_u':actual_gap_in_u,
               'channel_wt_per_m': channel_wt_per_m,
               'angle_wt_per_m' : angle_wt_per_m,
               'length_of_profiles': profile_length,
               'flange_height': flange_height,
               'width_of_profile': 0,
               'description': description }
       
    # -------- Channel & Angle Calculation -------- #
    material_totals = {}
    for crate_type, data in final_summary.items():
        num_crates = data["no_of_crates"]
        u_count = data["no_of_U_in_a_crate"]
        width = data["crate_width"]
        height = data["inner_crate_height"] + c_channel_height
        gap = data["gap_between_u"]
        total_height = data["crate_height"]
        angle_type = data["angle_weld_diagonal_or_box"]
        # C Channels required (3 lengths per crate)
        c1_len = width
        c1_qty = u_count * num_crates * 2
        c2_len = (height - 2 * c_channel_height) 
        c2_qty = u_count * num_crates * 2
        # Angles required
        if angle_type == 'Box':
            a_len = gap
            a_qty = (u_count - 1) * 4 * num_crates
        else :
            a_len = round(math.sqrt(gap ** 2 + total_height ** 2))
            a_qty = (u_count - 1) * 2 * num_crates
        # Save per-crate details
        data["Channel_requirements"] = {
            f"{c1_len}mm": c1_qty,
            f"{c2_len}mm": c2_qty }
    
        data["Angle_requirements"] = {
            f"{a_len}mm": a_qty    }
        # Add to total material list
        for l, q in data["Channel_requirements"].items():
            material_totals[f"Channel_{l}"] = material_totals.get(f"Channel_{l}", 0) + q
        for l, q in data["Angle_requirements"].items():
            material_totals[f"Angle_{l}"] = material_totals.get(f"Angle_{l}", 0) + q
    # Final material summary as DataFrame
    df = pd.DataFrame(material_totals.items(), columns=["Material", "Quantity (pcs)"])
    print("\n Material Requirement Summary:")
    print(df.to_string(index=False))
    
    # Initialize totals
    grand_total_channel_kg = 0
    grand_total_angle_kg = 0
    
    # Loop through crate types
    for crate_type, data in final_summary.items():
        channel_total_kg = 0
        angle_total_kg = 0
        # Calculate total channel weight
        for length_str, qty in data.get("Channel_requirements", {}).items():
            length_mm = int(float(length_str.replace("mm", "")))
            length_m = length_mm / 1000
            wt_per_m = data["channel_wt_per_m"]
            channel_total_kg += qty * length_m * wt_per_m
        # Calculate total angle weight
        for length_str, qty in data.get("Angle_requirements", {}).items():
            length_mm = int(float(length_str.replace("mm", "")))
            length_m = length_mm / 1000
            wt_per_m = data["angle_wt_per_m"]
            angle_total_kg += qty * length_m * wt_per_m
        # Add to dictionary
        data["total_channel_weight_kg"] = round(channel_total_kg, 2)
        data["total_angle_weight_kg"] = round(angle_total_kg, 2)
        # Add to grand totals
        grand_total_channel_kg += channel_total_kg
        grand_total_angle_kg += angle_total_kg
    # Final totals
    grand_total_channel_kg = round(grand_total_channel_kg, 2)
    grand_total_angle_kg = round(grand_total_angle_kg, 2)
    # Optional: show crate-wise dictionary
    print("\n Updated final Summary with Fabrication Details:")
    print(json.dumps(final_summary, indent=4))

def set_auto_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if cell.coordinate in sheet.merged_cells:
                continue
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                if not column_letter:
                    column_letter = cell.column_letter
        if column_letter:
            sheet.column_dimensions[column_letter].width = max_length + 2

def apply_border(cell):
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample BOM Dictionary
defined_bom = {
    "CE": {
        "elex": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 685, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61 },
            {"sl": 2, "name": "Top Hanger", "specs": "IS 2062", "length": 300, "width": 195, "thickness": 4, "pcs": 1, "weight": 1.84,"total_weight": 0,'rate_per_kg': 85},
            {"sl": 3, "name": "Bottom Hanger", "specs": "IS 2062", "length": 420, "width": 198, "thickness": 5, "pcs": 1, "weight": 3.26,"total_weight": 0,'rate_per_kg': 82},
            {"sl": 4, "name": "Rivet", "specs": "IS 1239", "length": 22, "width": 32, "thickness": 2.5, "pcs": 6, "weight": .24,"total_weight": 0,'rate_per_kg': 260},
            {"sl": 5, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},  
            {"sl": 6, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59},  
            {"sl": 7, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59}  ],
        "elex_cut": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 685, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "Top Hanger", "specs": "IS 2062", "length": 300, "width": 195, "thickness": 4, "pcs": 1, "weight": 1.84,"total_weight": 0,'rate_per_kg': 85},
            {"sl": 3, "name": "Cut Electrode Plate", "specs": "IS 2062", "length": 2095, "width": 395, "thickness": 5, "pcs": 1, "weight": 14.81,"total_weight": 0,'rate_per_kg': 80},
            {"sl": 4, "name": "MS Plate", "specs": "IS 2062", "length": 63, "width": 50, "thickness": 12, "pcs": 2, "weight": 0.59,"total_weight": 0,'rate_per_kg': 65},
            {"sl": 5, "name": "RIVET", "specs": "IS 1239", "length": 22, "width": 32, "thickness": 2.5, "pcs": 9, "weight": 0.36,"total_weight": 0,'rate_per_kg': 260},
            {"sl": 6, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},   
            {"sl": 7, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59},  
            {"sl": 8, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59}
            ],
        "elex_weld": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 685, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61 },
            {"sl": 2, "name": "Top Hanger", "specs": "IS 2062", "length": 300, "width": 195, "thickness": 3, "pcs": 1, "weight": 1.84,"total_weight": 0,'rate_per_kg': 85},
            {"sl": 3, "name": "Bottom Hanger", "specs": "IS 2062", "length": 420, "width": 198, "thickness": 4, "pcs": 1, "weight": 3.26,"total_weight": 0,'rate_per_kg': 82},
            {"sl": 5, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},  
            {"sl": 6, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59},  
            {"sl": 7, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 59}  ],
        "bhel": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 910, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "R. Clamp", "specs": "IS 513 CR2", "length": 60, "width": 90, "thickness": 1.6, "pcs": 2, "weight": 0.14,"total_weight": 0,'rate_per_kg': 110},
            {"sl": 3, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135}, 
            {"sl": 4, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58},  
            {"sl": 5, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58}
            ],
        "thermax": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 620, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},   
            {"sl": 3, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58},  
            {"sl": 4, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58}
            ],
        "thermax_497": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 620, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},  
            {"sl": 3, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58},  
            {"sl": 4, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58}
            ],
        "sitson": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 685, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},  
            {"sl": 3, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58},  
            {"sl": 4, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58}
            ],
        "bhel_387": [
            {"sl": 1, "name": "CRCA Sheet", "specs": "IS 513 CR2", "length": 1000, "width": 552, "thickness": 1.20, "pcs": 1, "weight": 0,"total_weight": 0,'rate_per_kg': 61},
            {"sl": 2, "name": "Suspension Hook", "specs": "IS 2062", "length": 218, "width": 100, "thickness": 5, "pcs": 1, "weight": 0.86,"total_weight": 0,'rate_per_kg': 96},
            {"sl": 3, "name": "Guide Iron", "specs": "IS 2062", "length": 85, "width": 40, "thickness": 3, "pcs": 1, "weight": 0.08,"total_weight": 0,'rate_per_kg': 115},
            {"sl": 4, "name": "Shock Iron", "specs": "IS 2062", "length": 265, "width": 150, "thickness": 6, "pcs": 1, "weight": 1.87,"total_weight": 0,'rate_per_kg': 65},
            {"sl": 5, "name": "RPO", "specs": "TA 506", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 1,"total_weight": 0,'rate_per_kg': 135},   
            {"sl": 6, "name": "Channel", "specs": "MS Steel", "length": 0, "width": 75, "thickness": 40, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58},  
            {"sl": 7, "name": "Angle", "specs": "MS Steel", "length": 0, "width": 0, "thickness": 0, "pcs": 0, "weight": 0,"total_weight": 0,'rate_per_kg': 58}
            ],    },
    "DE": {}      }         
wedge_type  = [{"type":1, "wd_length": 265, "wd_thickness": 3, "wd_width": 51, "wd_weight": .32, "rate_per_kg": 100,"length_cut":-70 },
               {"type":2, "wd_length": 195, "wd_thickness": 3, "wd_width": 30, "wd_weight": .14 ,"rate_per_kg": 100, "length_cut":-80}, 
               {"type":3, "wd_length": 195, "wd_thickness": 4, "wd_width": 30, "wd_weight": .185,"rate_per_kg": 100, "length_cut":-80},
               {"type":4, "wd_length": 106, "wd_thickness": 3, "wd_width": 100, "wd_weight": .251,"rate_per_kg": 100, "length_cut":-50},
               {"type":5, "wd_length": 106, "wd_thickness": 3, "wd_width": 150, "wd_weight": .374,"rate_per_kg": 100, "length_cut":-50},
               {"type":6, "wd_length": 139, "wd_thickness": 1.5, "wd_width":20, "wd_weight": .191,"rate_per_kg": 100, "length_cut":-73}  ]

nail_spike = [{"type":1,"name": "Nail","length":65, "width": 2.6,"thickness": 0, "rate_per_kg": 120,"nl_weight" : 0.00271, "specs": ""}, 
              {"type":2,"name": "Spike","length":47, "width": 15,"thickness": 1.2,"rate_per_kg": 120, "nl_weight" : 0.0066411, "specs": "IS 513 D"},
              {"type":3,"name": "Spike","length":46.65, "width": 15,"thickness": 1.25,"rate_per_kg": 120, "nl_weight" : 0.0068663,"specs": "IS 513 D" },
              {"type":4,"name": "Nail","length":65, "width": 2.7,"thickness": 0, "rate_per_kg": 120,"nl_weight" : 0.002923,"specs": ""},
              {"type":5,"name": "Nail","length":56, "width": 2.6,"thickness": 0,"rate_per_kg": 120, "nl_weight" : 0.002334,"specs": ""},
              {"type":6,"name": "Spike","length":60, "width": 24,"thickness":1, "rate_per_kg": 120,"nl_weight" : 0.011304,"specs": "IS 513 D"}  ]

components = [{"type":1,"name": "U Clamp","length":80, "width":35,"thickness": 2, "rate_per_kg": 80,"weight" : 0.151, "specs": "MS"}, 
              {"type":2,"name": "U Clamp","length":80, "width":35,"thickness": 1.6, "rate_per_kg": 80,"weight" : 0.121, "specs": "MS"},
              {"type":3,"name": "G D Screen Hook","length":232, "width":100,"thickness": 5, "rate_per_kg": 110,"weight" : 0.91, "specs": "MS"},
              {"type":4,"name": "Guide plate for CE","length":0, "width":0,"thickness": 2, "rate_per_kg": 100,"weight" : 1, "specs": "MS"},
              {"type":4,"name": "Top Tadpole","length":0, "width":0,"thickness": 4, "rate_per_kg": 150,"weight" : 1, "specs": "MS"},
              {"type":4,"name": "Bottom Tadpole","length":0, "width":0,"thickness": 3, "rate_per_kg": 150,"weight" : 1, "specs": "MS"},
              {"type":5,"name": "Holdng Plate","length":0, "width": 0,"thickness": 2,"rate_per_kg": 120, "weight" : 0.92, "specs": "MS"}  ]

def update_rate(data, new_rate,bom_name):
    for customer_group in data.values():
        for subcategory, items in customer_group.items():
            for item in items:
                if item.get("name") == bom_name:
                    item["rate_per_kg"] = new_rate

# Call the function
new_rate_crca = 61  #get_valid_float("Input New rate of CRCA: ",35,100)
bom_name = "CRCA Sheet"
update_rate(defined_bom, new_rate_crca,bom_name)  
new_rate_channel = 58    #get_valid_float("Input New rate of Channel: ",35,100)
bom_name = "Channel"
update_rate(defined_bom, new_rate_channel,bom_name)  
# new_rate_channel = get_valid_float("Input New rate of Angle: ",35,100)
bom_name = "Angle"
update_rate(defined_bom, new_rate_channel,bom_name) 
file_folder = "H:/My Drive/GDrive/Documents/test/"
# file_folder = "E/test"
# format_folder = "E:/test"
format_folder = "H:/My Drive/GDrive/Documents/test/"
file_path = f"{format_folder}/customer.xlsx"
sheet_name = "customers"
# Read with pandas for processing
df = pd.read_excel(file_path, sheet_name=sheet_name)
customers = df['short_name'].dropna().tolist()
short_name = choose_from_list("Select customer:", customers).strip().upper()
matched_row = df[df.iloc[:, 0].astype(str).str.strip().str.lower() == short_name.lower()]

# If match found, extract data
if not matched_row.empty:
    party = matched_row.iloc[0, 1]   # Column C (index 2)
    address_line1 = matched_row.iloc[0, 4]     # Column F (index 5)
    gst_number = matched_row.iloc[0, 5]      # Column G (index 6)
    city = matched_row.iloc[0, 6]
    contact_person = matched_row.iloc[0, 2]
    contact_no = matched_row.iloc[0, 3]
else:
    # Input new customer details
    print("Short name not found in the customer list, Enter details.")
    party = input("Enter full Name of Company: ")
    short_name = input("Enter Short Name of Company: ")
    contact_person = input("Enter Name of Person to Contact: ")
    contact_no = input("Enter Contact No of Person: ")
    address_line1 = input("Enter Company's full address: ")
    gst_number = input("Enter Company's GST No.: ")
    city = input("Enter City's Name of Customer': ")
    # Load the workbook and target sheet
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # Find the first completely empty row
    for row in range(2, sheet.max_row + 2):  # start from row 2 if row 1 is header
        if all(cell.value is None for cell in sheet[row]):
            last_blank_row = row
            break

    # Write the values to the blank row (adjust columns based on your format)
    sheet.cell(row=last_blank_row, column=1).value = short_name.upper()
    sheet.cell(row=last_blank_row, column=2).value = party
    sheet.cell(row=last_blank_row, column=3).value = contact_person
    sheet.cell(row=last_blank_row, column=4).value = contact_no
    sheet.cell(row=last_blank_row, column=5).value = address_line1
    sheet.cell(row=last_blank_row, column=6).value = gst_number
    sheet.cell(row=last_blank_row, column=7).value = city
    for col in range(1, 7):  # Adjust if you have more columns
        cell = sheet.cell(row=last_blank_row, column=col)
        apply_border(cell)
    # Save the file
    wb.save(file_path)
    wb.close()
    print(f"New customer '{party}' added to the Excel file.")
print(f"Selected customer: {party}")

file_path = f"{file_folder}/quote_list.xlsx" 
        
wb = openpyxl.load_workbook(file_path)
sheet = wb["25-26"]
last_row_f = sheet.max_row
while last_row_f > 0 and sheet.cell(row=last_row_f, column=6).value is None:
    last_row_f -= 1
cell_a = sheet.cell(row=last_row_f, column=1)
merged_ranges = sheet.merged_cells.ranges
last_quotation_no = None
for merged_range in merged_ranges:
    if cell_a.coordinate in merged_range:
        # Use the top-left cell of the merged range
        top_left = merged_range.min_row
        last_quotation_no = sheet.cell(row=top_left, column=1).value
        break
else:
    # Not part of merged range, use directly
    last_quotation_no = cell_a.value
try:
    quotation_number = int(last_quotation_no) + 1
except (TypeError, ValueError):
    print("Invalid quotation number in last row. Setting next number to 1.")
    quotation_number = 1
print(f"Last Quotation No: {last_quotation_no}")
print(f"Next Quotation No: {quotation_number}")
# print(f"Actual Last Row (based on column F): {last_row_f}")

quotation_date = datetime.today().strftime('%d-%b-%Y')
new_row = last_row_f + 1
sent_via = input("Reference no of enquiry received:  ")
enquiry_received_date = get_valid_date("Date of Enquiry received ((dd.mm.yy)): ")
quotation_folder_name = f"Quo{quotation_number}-{short_name}"
output_file =f"Cost-{quotation_folder_name}.xlsx"
if not output_file.lower().endswith(".xlsx"):
    output_file += ".xlsx"

# Folder path you want to create
# folder = f"{format_folder}/{quotation_folder_name}" 
folder = f"{file_folder}{quotation_folder_name}"
# Create the folder if it doesn't already exist
if not os.path.exists(folder):
    os.makedirs(folder)
    print(f"Folder '{folder}' created successfully.")
else:
    print(f"Folder '{folder}' already exists.")

sheet.cell(row=new_row, column=1).value = quotation_number
sheet.cell(row=new_row, column=2).value = quotation_date
sheet.cell(row=new_row, column=3).value = party
sheet.cell(row=new_row, column=4).value = city
sheet.cell(row=new_row, column=5).value = sent_via
sheet.cell(row=new_row, column=13).value = quotation_folder_name
start_row = new_row 
previous_u_gap = []
main_items = []
num_items = get_valid_int("Enter number of main items to include: ", 1 ,7)
for i in range(num_items):
    item_data = {}
    item_data["sl"] = i + 1
    item_data["item"] = input(f"Item {i+1} name (e.g., ce, de or x ): ").strip().upper()
    if item_data["item"] =="CE":
        valid_subtypes = list(defined_bom.get(item_data["item"], {}).keys())
        
        if not valid_subtypes:
            print(f"No sub-types found for item '{item_data['item']}'. Skipping.")
            continue  # or handle appropriately

        while True:
            # sub_type = input(f"Enter sub main item type for {item_data['item']} {valid_subtypes}: ").strip().lower()
            sub_type = choose_from_list("Select Design of CE: ", valid_subtypes).strip()
            if sub_type in valid_subtypes:
                break
            print(" Invalid sub main item type. Please enter one of:", valid_subtypes)

        if sub_type == "elex" or sub_type == "elex_cut" or sub_type == "elex_weld" or sub_type == "sitson":
            item_data["width_OD"] = 500
        elif sub_type == "bhel_387":
            item_data["width_OD"] = 387
        elif sub_type == "thermax_497":
            item_data["width_OD"] = 497    
        else: item_data["width_OD"] = float(input("Width Or OD (mm): ")) 
        item_data["desc"] = "Collecting Electrodes"
        item_data["thks"] = get_valid_float("Thickness (mm): (0.5 to 2.2): ", 0.5,2.2)
        item_data["t_length"] = get_valid_int("Total_Length including of attachments (mm): (2000 to 17000):  ", 2000, 17000)
        item_data["qty"] = get_valid_int(f"Item {i+1} quantity (Nos): (10 to 25000):  ", 10,25000)
        item_data["drawing"] = input(f"Item {i+1} Drawing No.: ")
        item_data["sell_price_kg"] = 100
        item_data['days_required'] = get_valid_float("No of Days required to complete the Order : ", 0.25, 200)
        weld_type = input("Enter angle weld type'diagonal'or'box': (d or b) ").strip().lower()
        if weld_type == 'b' :
           angle_weld_type = "Box" 
        else:  angle_weld_type = "Diagonal"
        profile_length = item_data["t_length"]
        num_profiles = item_data["qty"]
        profile_width = item_data["width_OD"]
        description = item_data["desc"]
        
        design_map = {
            'elex': "elex_pack",
            'elex_cut': "elexcut_pack",
            'elex_weld': "elex_nor_pack",
            'bhel': "bhel_pack",
            'bhel_387': "bhel387_pack",
            'thermax': "thermax20_pack",
            'sitson':"sitson_pack",
            'thermax_497': "thermax30_pack"   }   
        design = design_map.get(sub_type, None)  # or some default like `default_pack`
        product = item_data['item']
        nail_length = 0
        packing_details_def(product,profile_length, num_profiles, angle_weld_type, design,description,profile_width,nail_length)

        defined_bom["CE"][sub_type][0].update({"thickness":item_data["thks"]})
        defined_bom["CE"][sub_type][0].update({"length":item_data["t_length"]})
        if sub_type == "bhel" or sub_type == "thermax":
            defined_bom["CE"][sub_type][0].update({"width":(item_data["width_OD"]+ 175)})
        if sub_type == "elex" :
            defined_bom["CE"][sub_type][0].update({"length":(item_data["t_length"] - 220)})  
        if sub_type == "bhel_387" :
            defined_bom["CE"][sub_type][0].update({"length":(item_data["t_length"] - 195)})      
        if sub_type == "elex_cut" :
            defined_bom["CE"][sub_type][0].update({"length":(item_data["t_length"] - 1950)})  
        packing_weight = 0
        for bom_item in defined_bom['CE'][sub_type]:
            if bom_item['name'] == 'RPO':
                bom_item['weight'] = round((item_data['t_length'] * item_data['width_OD'] / 1000 / 1000 * 0.18),4)
                packing_weight = packing_weight + bom_item['weight']
        for bom_item in defined_bom['CE'][sub_type]:
            if bom_item['name'] == 'CRCA Sheet':
                bom_item['weight'] = round((bom_item['length'] * bom_item['width'] * bom_item['thickness'] / 1000000 * 7.85),4) 
        for bom_item in defined_bom['CE'][sub_type]:
            if bom_item['name'] == 'Channel':
                bom_item['weight'] =  round((grand_total_channel_kg / item_data['qty']),4)
                packing_weight = packing_weight + bom_item['weight']
        for bom_item in defined_bom['CE'][sub_type]:
            if bom_item['name'] == 'Angle':
                bom_item['weight'] =  round((grand_total_angle_kg / item_data['qty']),4) 
                packing_weight = packing_weight + bom_item['weight']
                
        item_data["bom"] = copy.deepcopy(defined_bom.get(item_data["item"], {}).get(sub_type, []))         
        x = 0
        for bom_item in item_data['bom']:
            bom_item['total_weight'] = round((bom_item['weight'] * item_data['qty']),4)
            x =  round((x + bom_item['weight']),4)
        for bom_item in item_data['bom']:
            bom_item['amount_per_pc'] = round((bom_item['weight'] * bom_item['rate_per_kg']),4)
        subtype = sub_type.upper()   
        main_items.append(item_data)
        main_items[i]['packing_weight']  = packing_weight
        main_items[i]["gross_wt_per_pc"] = x
        main_items[i]['total_net_wt'] = round(((x-packing_weight) * item_data["qty"]),2)
        main_items[i]['net_wt_per_pc'] = round((x-packing_weight),4)
        main_items[i]['sell_rate_per_pc'] = round(item_data["sell_price_kg"] * (x-packing_weight),2)
        main_items[i].setdefault('crate_details', {}).update(final_summary)
        main_items[i]['item_description'] = f"{main_items[i]['item']} ({subtype})"
    
         ###### DE starts ####
    elif item_data["item"] =="DE":
         while True:
             de_type = input(f"Item {i+1} Pipe (p), Strips (s) or Formed (f): ").strip().lower()
             if de_type in ['p', 's', 'f']:
                 break
             else:
                 print(f"Invalid input: '{de_type}'. Please enter 'p', 's', or 'f'.")
         # de_type = input(f"Item {i+1} Pipe (p), Strips (s) or Formed (f):  ")
         item_data["desc"] = "Discharge Electrodes"
         item_data["width_OD"] = get_valid_float("Width or OD (mm): ", 10, 200)
         item_data["thks"] = get_valid_float("Thickness (mm): ", 0.5, 3)
         item_data["t_length"] = get_valid_int("Total_Length including of attachments (mm): ", 1000, 17000)
         item_data["qty"] = get_valid_int(f"Item {i+1} quantity (Nos): ",10, 500000)
         item_data["drawing"] = input(f"Item {i+1} Drawing No.: ")
         item_data['sell_price_kg']   = 150
         # item_data["sell_rate_per_pc"] = get_valid_int("Selling rate per pc (Rs.): ", 100, 5000)
         pipe_rate_per_kg = get_valid_float("Rate per kg of Pipe or strip or : ", 40, 110)
         item_data['days_required'] = get_valid_float("No of Days required to complete the Order : ",0.25, 300)
         print("Available Wedge Types:")
         for wedge in wedge_type:
             print(f"Type {wedge['type']} - {wedge['wd_length']}x{wedge['wd_thickness']}x{wedge['wd_width']}")
         while True:
             attach_option = get_valid_int(
                 "\nAttach wedge? Enter 0 (No), 1 (One side), or 2 (Both sides): ",
                 0,
                 2,
             )
             if attach_option in [0, 1, 2]:
                 break
         length_cut = 0
         item_data.setdefault("bom", {})
         def get_wedge_type_input(label=""):
             while True:
                 try:
                     wedge_type_input = get_valid_int(
                         f"Enter wedge type to attach{label} (1 to 6): ", 1, 6
                     )
                     wedge = next((w for w in wedge_type if w['type'] == wedge_type_input), None)
                     if wedge:
                         return wedge
                     else:
                         print("Invalid type. Try again.")
                 except ValueError:
                     print("Enter a valid number.")
         if attach_option == 1:
              wedge = get_wedge_type_input(" (one side)")
              item_data["bom"]["wedge1"] = {
                  "type": wedge['type'],
                  "name": "Wedge1",
                  "length": wedge['wd_length'],
                  "thickness": wedge['wd_thickness'],
                  "width": wedge['wd_width'],
                  "pcs": 1,
                  "weight": wedge['wd_weight'],
                  "rate_per_kg" : wedge['rate_per_kg'],
                  "amount_per_pc": wedge['rate_per_kg'] * wedge['wd_weight'] ,
                  "total_weight" : wedge['wd_weight'] * item_data['qty'],
                  "specs" : "IS 3074",
                  "length_cut": wedge['length_cut']  }
              length_cut += wedge['length_cut']

         elif attach_option == 2:
              wedge1 = get_wedge_type_input(" (side A)")
              wedge2 = get_wedge_type_input(" (side B)")

              item_data["bom"]["wedge1"] = {
                  "type": wedge1['type'],
                  "name": "Wedge1",
                  "length": wedge1['wd_length'],
                  "thickness": wedge1['wd_thickness'],
                  "width": wedge1['wd_width'],
                  "pcs": 1,
                  "weight": wedge1['wd_weight'],
                  "rate_per_kg" : wedge1['rate_per_kg'],
                  "amount_per_pc": wedge1['rate_per_kg'] * wedge1['wd_weight'] ,
                  "total_weight" : wedge1['wd_weight'] * item_data['qty'],
                  "specs" : "IS 3074",
                  "length_cut": wedge1['length_cut']   }
          
              item_data["bom"]["wedge2"] = {
                  "type": wedge2['type'],
                  "name": "Wedge2",
                  "length": wedge2['wd_length'],
                  "thickness": wedge2['wd_thickness'],
                  "width": wedge2['wd_width'],
                  "pcs": 1,
                  "weight": wedge2['wd_weight'],
                  "rate_per_kg" : wedge2['rate_per_kg'],
                  "amount_per_pc": wedge2['rate_per_kg'] * wedge2['wd_weight'] ,
                  "total_weight" : wedge2['wd_weight'] * item_data['qty'],
                  "specs" : "IS 3074",
                  "length_cut": wedge2['length_cut'] }
              length_cut += wedge1['length_cut'] + wedge2['length_cut']                   
         
         item_data['bom']['pipe'] = {}  
         item_data['bom']['pipe']["length"] = item_data['t_length'] +  length_cut  
         item_data['bom']['pipe']["width"] = item_data['width_OD']
         item_data['bom']['pipe']['thickness'] = item_data['thks']
         item_data['bom']['pipe']['rate_per_kg'] = pipe_rate_per_kg
                
         if de_type == "p":
             joining_pcs = input("Joining pcs in pipe required? (y/n): ")
             item_data['bom']['pipe']['weight'] = round(3.1416 * item_data['bom']['pipe']["length"] * (item_data['width_OD'] - item_data['thks']) * item_data['thks'] * 7.85 / 1000000, 4)
             item_data['bom']['pipe']['specs'] = "IS 5429"
             item_data['bom']['pipe']["name"] = "Pipe"
             item_data['bom']['pipe']['pcs'] = 1 
         elif de_type =="f":
             item_data['bom']['pipe']["pipe_width_od"] = item_data['width_OD'] + 30
             item_data['bom']['pipe']['weight'] = (item_data['width_OD'] + 30) * item_data['thks'] * (item_data['t_length'] +  length_cut) * 2 * 7.85 /1000000
             item_data['bom']['pipe']['pcs'] = 2  
             item_data['bom']['pipe']['specs'] = "IS 513 D"
             item_data['bom']['pipe']["name"] = "CRCA Sheet"
         else:  
             item_data['bom']['pipe']['weight'] = (item_data['width_OD'] ) * item_data['thks'] * (item_data['t_length'] +  length_cut) *  7.85 /1000000   
             item_data['bom']['pipe']['specs'] = "IS 513 D"
             item_data['bom']['pipe']['pcs'] = 1
             item_data['bom']['pipe']["name"] = "CRCA Sheet"
         item_data['bom']['pipe']["total_weight"] = item_data['bom']['pipe']['weight'] * item_data['qty']
         item_data['bom']['pipe']["amount_per_pc"] = pipe_rate_per_kg * item_data['bom']['pipe']['weight']
         
         if  de_type == "p" and joining_pcs == "y":
             item_data['bom']['joining_pipe'] = {}
             item_data['bom']['joining_pipe']["name"] = "Joining Pipe"
             item_data['bom']['joining_pipe']["length"] = 200 
             item_data['bom']['joining_pipe']["width"] = item_data['width_OD']
             item_data['bom']['joining_pipe']["thickness"] = item_data['thks']
             if item_data['t_length'] > 12000:
                join_pc = 2
             else: join_pc = 1   
             item_data['bom']['joining_pipe']["pcs"] = join_pc
             item_data['bom']['joining_pipe']['specs'] = "IS 5429"
             item_data['bom']['joining_pipe']['weight'] = round(3.1416 * 150 * (item_data['width_OD'] - item_data['thks']) * item_data['thks'] * 7.85 / 1000000, 4)
             item_data['bom']['joining_pipe']["total_weight"] = item_data['bom']['joining_pipe']['weight'] * item_data['qty']
             item_data['bom']['joining_pipe']['rate_per_kg'] = pipe_rate_per_kg
             item_data['bom']['joining_pipe']["amount_per_pc"] = round(pipe_rate_per_kg * item_data['bom']['joining_pipe']['weight'],4)
             
         print("Available Spike / Nail Types:")
         for nail in nail_spike:
                 print(f"Type {nail['type']} {nail['name']} - {nail['width']} x {nail['thickness']} x {nail['length']}")
         nail_type_input = get_valid_int(
             "Enter Spike or Nail type to attach (1 to 6): ", 1, 6
         )
         spike_qty = get_valid_float("Enter Spike or Nail quantity required per pc: ",  10, 300)
         spike_nail = next((w for w in nail_spike if w['type'] == nail_type_input), None) 
         spike_nail["pcs"] = spike_qty  
         spike_nail["weight"] = spike_qty  * spike_nail['nl_weight']
         spike_nail["total_weight"]  = round(spike_nail["weight"] * item_data['qty'],4)
         spike_nail["amount_per_pc"] = round(spike_nail["weight"] * spike_nail['rate_per_kg'],4)
         item_data["bom"]["spike_nails"] = spike_nail
         net_wt = 0
         for component in item_data['bom'].values():
             wt = component.get('weight', 0)
             net_wt += wt
         item_data['sell_rate_per_pc']   = round(item_data['sell_price_kg'] * net_wt, 2)
         item_data['net_wt_per_pc'] = net_wt
         item_data['total_net_wt'] = round(net_wt *  item_data["qty"],2)  
         
         profile_length = item_data["t_length"]
         num_profiles = item_data["qty"]
         weld_type = input("Enter angle weld type'diagonal'or'box': (d or b) ").strip().lower()
         if weld_type == 'd' or weld_type == "":
            angle_weld_type = "Diagonal"
         else: angle_weld_type = "Box" 
         design = "std_pack"
         product = item_data['item']
         description = item_data['desc']
         profile_width = item_data['width_OD']
         nail_length = spike_nail['length']
         packing_details_def(product,profile_length, num_profiles, angle_weld_type, design,description,profile_width, nail_length)
         bom_dict = item_data.get('bom', {})
         bom_dict['RPO'] = {key: 0 for key in bom_dict['pipe']}
         bom_dict['RPO']['weight'] = round((profile_width * profile_length/1000/1000*0.16*3.15)*2,4)
         bom_dict['RPO']['total_weight'] = bom_dict['RPO']['weight'] * num_profiles
         bom_dict['RPO']['specs']  = "TA 506"
         bom_dict['RPO']['rate_per_kg'] = 135
         bom_dict['RPO']['name'] = "RPO"
         bom_dict['RPO']['amount_per_pc'] = bom_dict['RPO']['weight'] * 135
         bom_dict['channel'] = {key: 0 for key in bom_dict['pipe']}
         bom_dict['channel']['weight'] =round(grand_total_channel_kg / num_profiles,4)
         bom_dict['channel']['total_weight'] = grand_total_channel_kg
         bom_dict['channel']['specs']  = "MS Steel"
         rate_channel_angle = defined_bom['CE']['elex'][5]['rate_per_kg']
         bom_dict['channel']['rate_per_kg'] = rate_channel_angle
         bom_dict['channel']['name'] = "Channel"
         bom_dict['channel']['amount_per_pc'] = round(bom_dict['channel']['weight'] * rate_channel_angle,4)
         bom_dict['angle'] = {key: 0 for key in bom_dict['pipe']}
         bom_dict['angle']['weight'] =round(grand_total_angle_kg / num_profiles,4)
         bom_dict['angle']['total_weight'] = grand_total_angle_kg
         bom_dict['angle']['specs']  = "MS Steel"
         bom_dict['angle']['rate_per_kg'] = rate_channel_angle
         bom_dict['angle']['name'] = "Angle"
         bom_dict['angle']['amount_per_pc'] = round(bom_dict['angle']['weight'] * rate_channel_angle,4)
         bom_list = list(bom_dict.values())
         # Reassign sl values based on rules
         new_bom_list = []
         serial = 3
         for part in bom_list:
             name = part.get("name", "").strip().lower()
             if name in ["pipe", "crca sheet"]:
                 part["sl"] = 1
             elif name in ["nail", "spike"]:
                part["sl"] = 2
             else:
                 part["sl"] = serial
                 serial += 1
             new_bom_list.append(part)
         new_bom_list.sort(key=lambda x: x.get("sl", 999))
         # Now assign the list back
         item_data["bom"] = new_bom_list
         x = 0
         for bom_item in item_data['bom']:
             bom_item['total_weight'] = round((bom_item['weight'] * item_data['qty']),4)
             x =  round((x + bom_item['weight']),4)
         item_data["gross_wt_per_pc"] = x    
         main_items.append(item_data) 
         main_items[i]['item_description'] = f"{main_items[i]['item']} "
         main_items[i].setdefault('crate_details', {}).update(final_summary)
    else: 
         # components = item_data['item']
         for comps in components:
                 print(f"Type {comps['type']} {comps['name']} - {comps['width']} x {comps['thickness']} x {comps['length']}")
         comp_type_input = get_valid_int("Select Component - type (1,2,3,...): ", 1, 5)
         comp_qty = get_valid_int("Enter Component quantity required :  ",1,500000)
         # sell_rate = get_valid_int("Enter Selling price per pc :  ",5,500000)
         comp_detail = next((w for w in components if w['type'] == comp_type_input), None) 
         name = comp_detail['name']
         weight = comp_detail['weight']
         length = comp_detail['length']
         width = comp_detail['width']
         thickness = comp_detail['thickness']
         rate_per_kg = comp_detail['rate_per_kg']
         rate_per_pc = rate_per_kg * weight
         total_weight =  weight * comp_qty
         specs = comp_detail['specs']
         sell_price_per_kg = 150
         sell_rate = sell_price_per_kg * weight
         margin_per_pc = sell_rate - rate_per_pc
         margin_per_kg = sell_price_per_kg - rate_per_kg
         sl = i + 1
         main_items.append(  
            { 'sl': sl,
              'item': name,
              'width_OD': width,
              'desc': name,
              'thks': thickness,
              't_length': length,
              'qty': comp_qty,
              'drawing': '',
              'sell_rate_per_pc': sell_rate,
              'days_required': 1,
              'bom': [
                  {   'sl': 1,
                      'name': name,
                      'specs': specs,
                      'length':length ,
                      'width':width ,
                      'thickness': thickness,
                      'pcs': comp_qty,
                      'weight': weight,
                      'total_weight': total_weight,
                      'rate_per_kg': rate_per_kg,
                      'amount_per_pc': rate_per_pc,
                      'type': 0,  # optional key for some BOM entries
                      'length_cut': 0    }  ],
              'packing_weight': 0.0,
              'gross_wt_per_pc':weight,
              'total_net_wt': total_weight,
              'net_wt_per_pc': weight,
              'sell_price_kg': sell_price_per_kg,
              'crate_details': {
                  'Crate Type 1': {
                      'no_of_crates': 1,
                      'profiles_per_crate': comp_qty,
                      'crate_width': 0.0,
                      'crate_length': 0.0,
                      'crate_height': 0.0,
                      'inner_crate_height': 0.0,
                      'no_of_U_in_a_crate': 0,
                      'angle_weld_diagonal_or_box': '',
                      'gap_between_u': 0.0,
                      'channel_wt_per_m': 0.0,
                      'angle_wt_per_m': 0.0,
                      'length_of_profiles': 0.0,
                      'flange_height': 0.0,
                      'width_of_profile': 0.0,
                      'description': name,
                      'Channel_requirements': {
                          'key': 0       },
                      'Angle_requirements': {
                          'key': 0    },
                      'total_channel_weight_kg': 0.0,
                      'total_angle_weight_kg': 0.0,
                      'net_weight_per_crate': 0.0,
                      'gross_weight_per_crate': 0.0       },
              },
              'item_description': name,
              'cost_per_pc': rate_per_pc,
              'cost_per_kg': rate_per_kg,
              'margin_per_kg': margin_per_kg,
              'margin_per_pc': margin_per_pc,
              'production_per_day_pc': 1,
              'margin_per_day': margin_per_pc * comp_qty       }      )

#------------------------------------------------------------------------- Quote List Starts -----------------------------------------------------------------

k = 0
for item in main_items:
    bom_list = item.get('bom', [])
    total_material_cost = sum(bom.get('amount_per_pc', 0) for bom in bom_list)
    item['total_material_cost_per_pc'] = round(total_material_cost,0)
    while True:
        try:
            rate_offered_per_kg = get_valid_float(f"Price offered per kg of item No. {k+1} :  ", 60 , 250)
            rate_offered_per_pc = round(rate_offered_per_kg * item['net_wt_per_pc'],0)
            # margin = calculate_gross_margin(rate_offered, total_material_cost)
            gross_margin_per_kg = round((rate_offered_per_pc - total_material_cost) / (item['net_wt_per_pc']),2)
            print(f"Expected Gross Margin (Includes material & Packing cost) per kg of item No.{k+1}: {gross_margin_per_kg} ")
            satisfied = input("Offered price is comfortable? (y / n): ").strip().lower()
            if satisfied in ['yes', 'y']:
                # print("Loop exited.")
                break
        except ValueError:
            print("Please enter a valid numeric rate.")
    item['gross_margin_per_kg'] = gross_margin_per_kg
    item['rate_offered_per_kg']= rate_offered_per_kg
    item['rate_offered_per_pc']= rate_offered_per_pc
    new_row = k + start_row
    sheet.cell(row=new_row, column=6).value = item['item_description'] #item_des
    sheet.cell(row=new_row, column=7).value = item['width_OD']#profile_width
    sheet.cell(row=new_row, column=8).value = item['thks']
    sheet.cell(row=new_row, column=9).value = item['t_length']
    sheet.cell(row=new_row, column=10).value = item['qty']
    sheet.cell(row=new_row, column=11).value = item['net_wt_per_pc']
    sheet.cell(row=new_row, column=14).value = item['rate_offered_per_pc']
    sheet.cell(row=new_row, column=15).value = item['rate_offered_per_kg']
    for col_letter in "ABCDEFGHIJKLMNOPQRS":
        col_num = column_index_from_string(col_letter)
        cell = sheet.cell(row=new_row, column=col_num)
        apply_border(cell)
    k += 1

end_row = start_row + i   
columns_to_merge = [1, 2, 3, 4, 5, 13]   
for col in columns_to_merge:
    sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    
end_row = start_row + k - 1   
columns_to_merge = [1, 2, 3, 4, 5]   
for i in range(len(main_items)):
   for col in columns_to_merge:
       sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
   if quotation_number % 2 == 0:
       fill_color = "B0C4DE"  # Blue
   else:
       fill_color = "FFCCCC"  # Light Red
   fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
   for col in range(1, 21):  
       cell = sheet.cell(row=start_row + i, column=col)
       cell.fill = fill   
wb.save(file_path)
wb.close()
print(f"New quotation '{party}' added to the Excel file.")
#--------------------------------------------------------------------- Quote List Ends ---------------------------------------------------------------------
     
for item in main_items:
    wt_per_pc = item.get('net_wt_per_pc', 0)  # default 0 if not present
    gr_wt_pc = item.get('gross_wt_per_pc', 0) 
    crate_details = item.get('crate_details', {})

    for crate in crate_details.values():  # loop through Crate Type 1, 2, etc.
        no_profies = crate.get('profiles_per_crate')
        crate['net_weight_per_crate'] = round(wt_per_pc * no_profies,2)
        crate['gross_weight_per_crate'] = round(gr_wt_pc * no_profies,2)
all_crate_data = []
for item in main_items:
    item_id = item.get("sl", "NA")
    item_name = item.get("item", "NA")
    crate_details = item.get("crate_details", {})
    for crate_type, details in crate_details.items():
        # Flatten nested dictionary into a single row
        flat_data = {
            "Item SL": item_id,
            "Item Name": item_name,
            "Crate Type": crate_type,        }
        flat_data.update(details)
        all_crate_data.append(flat_data)
n = 1
for j in all_crate_data:
    j['Crate Type'] = f"Crate type {n}"
    n +=1 
    
max_vehicle_bed = get_valid_int("Enter Tralier's bed width available': ", 1400, 2700)     
max_height_trailer = get_valid_int("Enter Tralier's height available': ",1500,3000)      
crate_between_space_in_width = 100 
crate_between_space_in_height = 30           
# - ---------------------------------------------------------------------- COST SHEET PRINT STARTS ----------------------------------------------------------        
      
# Write to Excel
wb = Workbook()
sheet = wb.active
sheet.title = "Costing"
sheet.merge_cells("A1:I1")
sheet["A1"] = "LIRCO ENGINEERING PVT LTD"
sheet["A1"].font = Font(bold=True, size=16)
sheet["A1"].alignment = Alignment(horizontal="center")
sheet.merge_cells("A3:I3")
sheet["A3"] = "COST CALCULATION"
sheet["A3"].font = Font(bold=True, size=14)
sheet["A3"].alignment = Alignment(horizontal="center")
sheet.merge_cells("A5:C5")
sheet["A5"] = f"Quotation No. - {quotation_folder_name}"
sheet.merge_cells("A6:C6")
sheet["A6"] = f"Quotation Date. - {quotation_date}"
sheet.merge_cells("A7:C7")
sheet["A7"] = f"Customer Name. - {party}"
sheet.merge_cells("H5:I5")
sheet["H5"] = f"Reference  - {sent_via}"
# sheet.merge_cells("H6:I6")
# sheet["H6"] = f"Projected Start Date - {start_date}"
# sheet.merge_cells("H7:I7")
# sheet["H7"] = f"Delivery Date - {delivery_date}"

main_start_row = 9
headers = ["Sl. No.", "Item", "Description", "Width Or OD","Thickness","Total_Length","Quantity (Nos.)", "Total Net Weight (Kg.)","offered Price per pc", "Material Cost per PC.","Gross Margin Per Kg"  ]
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=main_start_row, column=col, value=header)
    cell.font = Font(bold=True)
    apply_border(cell)

current_row = main_start_row + 1
for item in main_items:
    values = [item["sl"], item["item"], item["desc"], item["width_OD"], item["thks"], item["t_length"], item["qty"], item["total_net_wt"],item["rate_offered_per_pc"],item["total_material_cost_per_pc"],item["gross_margin_per_kg"]]
    for col, val in enumerate(values, 1):
        cell = sheet.cell(row=current_row, column=col, value=val)
        apply_border(cell)
    current_row += 1

# if print_work_order =='y':
current_row += 2
sheet.merge_cells(f"A{current_row}:I{current_row}")
sheet[f"A{current_row}"] = "RAW MATERIAL DETAILS WITH COST CALCULATION"
sheet[f"A{current_row}"].font = Font(bold=True, size=12)
sheet[f"A{current_row}"].alignment = Alignment(horizontal="center")
current_row += 1

for item in main_items:
    sheet.merge_cells(f"A{current_row}:D{current_row}")
    sheet[f"A{current_row}"] = f"LIST OF MATERIALS FOR: {item['item']}"
    sheet[f"A{current_row}"].font = Font(bold=True)
    current_row += 1

    bom_headers = ["Sl. No.", "Name", "Specs",  "Width (mm.)", "Thk. (mm.)",  "Length (mm.)", "Qty (Nos.)", "Weight (Kg./pc)","Total Weight","Rate per Kg.","Amount per Pc"]
    for col, header in enumerate(bom_headers, 1):
        cell = sheet.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        apply_border(cell)
    current_row += 1   
    cost_per_pc = 0
    for bom in item["bom"]:
        row = current_row 
        cost_per_pc += bom.get('amount_per_pc', 0)
        values = [bom["sl"], bom["name"], bom["specs"],  bom["width"], bom["thickness"],bom["length"], bom["pcs"], bom["weight"],bom["total_weight"],bom["rate_per_kg"],bom["amount_per_pc"]]
        for col, val in enumerate(values, 1):
            cell = sheet.cell(row=row, column=col, value=val)
            apply_border(cell)
        current_row = row + 1
    item['cost_per_pc'] =  cost_per_pc 
    item['cost_per_kg'] =  round(cost_per_pc / item['net_wt_per_pc'],4)
    item['margin_per_kg'] = round(item['rate_offered_per_kg'] - cost_per_pc / item['net_wt_per_pc'],4)
    item['margin_per_pc'] = round((item['rate_offered_per_pc'] - cost_per_pc),4)
    item['production_per_day_pc'] = item['qty'] / item['days_required']
    item['margin_per_day'] = round(item['margin_per_pc'] * item['production_per_day_pc'],4)
    
    sheet[f"G{current_row}"] = "Total"
    sheet[f"H{current_row}"] = f"=SUM(H{current_row - len(item['bom'])}:H{current_row - 1})"
    sheet[f"I{current_row}"] = f"=SUM(I{current_row - len(item['bom'])}:I{current_row - 1})"
    sheet[f"K{current_row}"] = f"=SUM(K{current_row - len(item['bom'])}:K{current_row - 1})"
    
    apply_border(sheet[f"G{current_row}"])
    apply_border(sheet[f"H{current_row}"])
    apply_border(sheet[f"I{current_row}"])
    apply_border(sheet[f"J{current_row}"])
    apply_border(sheet[f"K{current_row}"])
    current_row += 1

sheet.merge_cells(f"A{current_row}:C{current_row}")
sheet[f"A{current_row}"] = "EXPECTED CRATE DETAILS"  
sheet[f"A{current_row}"].font = Font(bold=True)
current_row += 1 
crate_headers = ["Crate Type ", "Item & No of Crates",  "Crate Width x Height x Length (mm.)", "Angle Diagonal or Box", "No of U per crate & Gap between U","Profiles per crate","Total channel No. (Length)","Total channel No. (Length)","Total Angle No. (Length)"]
for col in range(1, 10):  # Columns A (1) to I (9)
    cell = sheet.cell(row = current_row, column=col)
    cell.alignment = Alignment(wrap_text=True)
sheet.row_dimensions[current_row].height = 48    
for col, header in enumerate(crate_headers, 1):
    cell = sheet.cell(row=current_row, column=col, value=header)
    cell.font = Font(bold=True)
    apply_border(cell)
current_row += 1 

for lists in all_crate_data:
   if lists['Item Name'] =='CE' or lists['Item Name'] =='DE':
      named = {}
      namea = {}
      for i, (size, qty) in enumerate(lists['Channel_requirements'].items(), 1):
          named[f"{i}"] = [size,qty]
      
      for i, (size, qty) in enumerate(lists['Angle_requirements'].items(), 1):
          namea[f"{i}"] = [size,qty]    
      row = current_row 
      values = [f"{lists['Crate Type']}  ",
                f"{lists['Item Name']} ({lists['no_of_crates']} Nos.)",
                f"{round(lists['crate_width'],0)} x {round(lists['crate_height'],0)} x {round(lists['crate_length'],0)}",
                f"{lists['angle_weld_diagonal_or_box']} ",
                f"{lists['no_of_U_in_a_crate']}Nos. ({lists['gap_between_u']}mm)",
                f"{lists['profiles_per_crate']} Nos. ",
                f"{named['1'][0]} - {named['1'][1]}pcs",
                f"{named['2'][0]} - {named['2'][1]}pcs",
                f"{namea['1'][0]} - {namea['1'][1]}pcs"   ]
      for col, val in enumerate(values, 1):
          cell = sheet.cell(row=row, column=col, value=val)
          apply_border(cell)
      current_row = row + 1
for col in range(1, 10):  # Columns A to I = 1 to 9
    col_letter = openpyxl.utils.get_column_letter(col)
    sheet.column_dimensions[col_letter].width = 11   
# set_auto_width(sheet)
full_paths = os.path.join(folder, output_file)
# Create a workbook and save to the path
wb.save(full_paths)
print(f" File saved to: {full_paths}")
# - --------------------------------------------------------------------------Cost sheet  PRINTING ENDS ----------------------------------------------------------        
        
packing_list = {}
packing_list["quotation_no"] = quotation_folder_name   
packing_list["pl_no"] = "Provisinal"
packing_list["address"] = address_line1
packing_list["contact_person"] = contact_person
packing_list["contact_no"] = contact_no
packing_list["quotation_date"] = quotation_date
packing_list["enquiry_date"] = enquiry_received_date
packing_list["enquiry_reference"] = sent_via
packing_list["max_vehicle_bed"] = max_vehicle_bed
packing_list["max_height_trailer"] = max_height_trailer
packing_list["crate_between_space_in_width"] = crate_between_space_in_width
packing_list["crate_between_space_in_height"] = crate_between_space_in_height
packing_list["party"] = party 
packing_list['packing_details'] = copy.deepcopy(all_crate_data)

keys_to_delete = ["Angle_requirements", "angle_weld_diagonal_or_box",'angle_wt_per_m','Channel_requirements','channel_wt_per_m',
                  'inner_crate_height','crate_length','total_channel_weight_kg','total_angle_weight_kg','no_of_U_in_a_crate','gap_between_u']
for item in packing_list['packing_details']:
    item['total_packing_weight'] = round(item['total_angle_weight_kg'] + item['total_channel_weight_kg'],0)
    item['total_net_weight'] = round(item['net_weight_per_crate'] * item['no_of_crates'] ,0)
    item['total_gross_weight'] = round(item['gross_weight_per_crate'] * item['no_of_crates'] ,0)
    item['total_pcs'] = item['profiles_per_crate'] * item['no_of_crates']
    item['length_display'] = item.get('length_of_profiles', 0) + 100
    width_for_height = item.get('width_of_profile', 0) or 0
    item['crate_height_display'] = (
        width_for_height + 120 if width_for_height else item.get('crate_height', 0)
    )
    for key in keys_to_delete:
        item.pop(key, None) 

def calculate_trailer_loading_separately_by_length(packing_list):
    max_width = packing_list['max_vehicle_bed']
    max_height = packing_list['max_height_trailer']
    space_w = packing_list['crate_between_space_in_width']
    space_h = packing_list['crate_between_space_in_height']
    profile_length_map = defaultdict(list)
    for item in packing_list['packing_details']:
        profile_length = item['length_of_profiles']
        gross_weight = item['gross_weight_per_crate']
        for _ in range(item['no_of_crates']):
            crate = {
                'crate_type': f"Crate Type {item['Crate Type'].split()[-1]}",
                'width': item['crate_width'],
                'height': item.get('crate_height_display', item['crate_height']),
                'length': item.get('length_display', item['length_of_profiles']),
                'gross_weight': gross_weight
            }
            profile_length_map[profile_length].append(crate)

    all_trailers = []
    for length_key, crates in profile_length_map.items():
        crates.sort(key=lambda x: x['width'], reverse=True)
        trailers = []
        current_rows, row = [], []
        row_width = row_max_height = used_height = max_trailer_width_used = 0
        def finalize_trailer(rows_data, max_width_used):
            trailer_summary = defaultdict(int)
            trailer_weight = height_used = 0
            for idx, row in enumerate(rows_data):
                row_height = max(crate['height'] for crate in row)
                if idx > 0:
                    height_used += space_h
                height_used += row_height
                for crate in row:
                    trailer_summary[crate['crate_type']] += 1
                    trailer_weight += crate['gross_weight']
            trailer_summary['Max Width Used (mm)'] = max_width_used
            trailer_summary['Max Height Used (mm)'] = height_used
            trailer_summary['Total Gross Weight (kg)'] = trailer_weight
            trailer_summary['Max length Used (mm)'] = length_key + 100
            trailer_summary['Profile Length (mm)'] = length_key + 100
            return trailer_summary
        for crate in crates:
            w, h = crate['width'], crate['height']
            if row_width + w + (len(row) * space_w) <= max_width:
                row.append(crate)
                row_width += w
                row_max_height = max(row_max_height, h)
            else:
                if used_height + row_max_height + (space_h if current_rows else 0) > max_height:
                    trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
                    current_rows, used_height, max_trailer_width_used = [], 0, 0
                if current_rows:
                    used_height += space_h
                used_height += row_max_height
                current_rows.append(row)
                max_trailer_width_used = max(max_trailer_width_used, row_width + (len(row) - 1) * space_w)
                row = [crate]
                row_width = w
                row_max_height = h
        if row:
            if used_height + row_max_height + (space_h if current_rows else 0) > max_height:
                trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
                current_rows = []
            if current_rows:
                used_height += space_h
            used_height += row_max_height
            current_rows.append(row)
            max_trailer_width_used = max(max_trailer_width_used, row_width + (len(row) - 1) * space_w)
        if current_rows:
            trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
        all_trailers.extend(trailers)
    return {f"Trailer {i+1}": dict(t) for i, t in enumerate(all_trailers)}

def calculate_trailer_loading(packing_list):
    max_width = packing_list['max_vehicle_bed']
    max_height = packing_list['max_height_trailer']
    space_w = packing_list['crate_between_space_in_width']
    space_h = packing_list['crate_between_space_in_height']
    all_crates = []
    crate_type_id = 1
    for item in packing_list['packing_details']:
        gross_weight = item['gross_weight_per_crate']
        for _ in range(item['no_of_crates']):
            crate = {
                'crate_type': f"Crate Type {crate_type_id}",
                'width': item['crate_width'],
                'height': item.get('crate_height_display', item['crate_height']),
                'length': item.get('length_display', item['length_of_profiles']),
                'gross_weight': gross_weight     }
            all_crates.append(crate)
        crate_type_id += 1
    all_crates.sort(key=lambda x: x['width'], reverse=True)
    trailers = []
    current_rows = []
    row = []
    row_width = 0
    row_max_height = 0
    used_height = 0
    max_trailer_width_used = 0
    def finalize_trailer(rows_data, max_width_used):
        trailer_summary = defaultdict(int)
        trailer_weight = 0
        height_used = 0
        max_length_used = 0
        for idx, row in enumerate(rows_data):
            row_height = max(crate['height'] for crate in row)
            row_length = max(crate['length'] for crate in row)
            if idx > 0:
                height_used += space_h
            height_used += row_height
            max_length_used = max(max_length_used, row_length)
            for crate in row:
                trailer_summary[crate['crate_type']] += 1
                trailer_weight += crate['gross_weight']
        trailer_summary['Max Width Used (mm)'] = max_width_used
        trailer_summary['Max Height Used (mm)'] = height_used
        trailer_summary['Max length Used (mm)'] = max_length_used
        trailer_summary['Total Gross Weight (kg)'] = trailer_weight
        return trailer_summary

    for crate in all_crates:
        w, h = crate['width'], crate['height']
        if row_width + w + (len(row) * space_w) <= max_width:
            row.append(crate)
            row_width += w
            row_max_height = max(row_max_height, h)
        else:
            if used_height + row_max_height + (space_h if current_rows else 0) > max_height:
                trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
                current_rows = []
                used_height = 0
                max_trailer_width_used = 0
            if current_rows:
                used_height += space_h
            used_height += row_max_height
            current_rows.append(row)
            max_trailer_width_used = max(max_trailer_width_used, row_width + (len(row) - 1) * space_w)
            row = [crate]
            row_width = w
            row_max_height = h

    if row:
        if used_height + row_max_height + (space_h if current_rows else 0) > max_height:
            trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
            current_rows = []
            used_height = 0
            max_trailer_width_used = 0
        if current_rows:
            used_height += space_h
        used_height += row_max_height
        current_rows.append(row)
        max_trailer_width_used = max(max_trailer_width_used, row_width + (len(row) - 1) * space_w)

    if current_rows:
        trailers.append(finalize_trailer(current_rows, max_trailer_width_used))
    trailer_result = {f"Trailer {i+1}": dict(t) for i, t in enumerate(trailers)}
    # Export to Excel
    trailer_df_list = []
    for trailer_id, content in trailer_result.items():
        row_data = {'Trailer': trailer_id}
        row_data.update(content)
        trailer_df_list.append(row_data)
    return trailer_result

def summarize_loading_types(trailer_result):
    loading_type_map = {}
    crate_combination_to_type = {}
    current_type_id = 1
    for trailer_id, trailer_data in trailer_result.items():
        crate_counts = {k: v for k, v in trailer_data.items() if k.startswith("Crate Type")}
        max_width = trailer_data.get("Max Width Used (mm)", 0)
        max_height = trailer_data.get("Max Height Used (mm)", 0)
        total_weight = trailer_data.get("Total Gross Weight (kg)", 0)
        crate_tuple = tuple(sorted(crate_counts.items()))
        if crate_tuple not in crate_combination_to_type:
            crate_combination_to_type[crate_tuple] = f"Loading Type {current_type_id}"
            current_type_id += 1
        loading_type = crate_combination_to_type[crate_tuple]
        if loading_type not in loading_type_map:
            loading_type_map[loading_type] = {
                'trailers': [],
                'crate_counts': crate_counts,
                'max_width': max_width,
                'max_height': max_height,
                'total_weight': total_weight            }
        else:
            loading_type_map[loading_type]['max_width'] = max(
                loading_type_map[loading_type]['max_width'], max_width)
            loading_type_map[loading_type]['max_height'] = max(
                loading_type_map[loading_type]['max_height'], max_height)
            loading_type_map[loading_type]['total_weight'] += total_weight
        loading_type_map[loading_type]['trailers'].append(trailer_id)
    return loading_type_map

def get_loading_type_summary(trailer_result):
    loading_type_map = {}
    crate_combination_to_type = {}
    current_type_id = 1
    for trailer_id, trailer_data in trailer_result.items():
        crate_counts = {k: v for k, v in trailer_data.items() if k.startswith("Crate Type")}
        max_width = trailer_data.get("Max Width Used (mm)", 0)
        max_height = trailer_data.get("Max Height Used (mm)", 0)
        total_weight = trailer_data.get("Total Gross Weight (kg)", 0)
        max_length = trailer_data.get("Max length Used (mm)", 0)
        crate_tuple = tuple(sorted(crate_counts.items()))
        if crate_tuple not in crate_combination_to_type:
            crate_combination_to_type[crate_tuple] = f"Loading Type {current_type_id}"
            current_type_id += 1
        loading_type = crate_combination_to_type[crate_tuple]
        if loading_type not in loading_type_map:
            loading_type_map[loading_type] = {
                'trailers': [],
                'crate_counts': crate_counts,
                'max_width': max_width,
                'max_height': max_height,
                'max_length': max_length,
                'total_weight': total_weight            }
        else:
            loading_type_map[loading_type]['max_width'] = max(
                loading_type_map[loading_type]['max_width'], max_width)
            loading_type_map[loading_type]['max_height'] = max(
                loading_type_map[loading_type]['max_height'], max_height)
            loading_type_map[loading_type]['max_length'] = max_length
            loading_type_map[loading_type]['total_weight'] += total_weight
        loading_type_map[loading_type]['trailers'].append(trailer_id)
    # Build final structured dictionary Trailer
    loading_summary = {}
    for loading_type, info in loading_type_map.items():
        crate_summary = ', '.join([f"{ctype} {qty} nos" for ctype, qty in info['crate_counts'].items()])
        loading_summary[loading_type] = {
            "Used_in_trailers": len(info['trailers']),
            "Crate_Summary": crate_summary,
            "Max_Width_mm": int(info['max_width']),
            "Max_Height_mm": int(info['max_height']),
            "Max_Length_mm": int(info['max_length']),
            "Gross_Weight_per_trailer_kg": info['total_weight']  /   len(info['trailers'])    }
    return loading_summary

group_by_length = input("Do you want to load crates in separate trailers based on profile length? (y/n): ").strip().lower()
if group_by_length == 'y':
    trailer_result = calculate_trailer_loading_separately_by_length(packing_list)
else:
    trailer_result = calculate_trailer_loading(packing_list)
loading_type_summary = get_loading_type_summary(trailer_result)
packing_list['trailer_loading_details'] = loading_type_summary.copy()
summarize_loading_types(trailer_result)
# - --------------------------------------------------------------------------PACKING LIST PRINT STARTS ----------------------------------------------------------        
# Template and new file
template_file = f"{format_folder}/Provisional_Packing_List.xlsx"
new_file = f"PL-{packing_list['quotation_no']}.xlsx"
# Load workbook
wb = openpyxl.load_workbook(template_file)
ws = wb.active
# Border and font
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
bold_font = Font(bold=True)

# Helper for safe write to merged cells
def safe_write(ws, cell, value):
    top_lefts = [rng.coord.split(":")[0] for rng in ws.merged_cells.ranges]
    if cell in top_lefts or cell not in [rng.coord.split(":")[1] for rng in ws.merged_cells.ranges]:
        ws[cell].value = value
    else:
        print(f" Skipped writing to merged cell {cell} (not top-left).")
# Fill header fields
safe_write(ws, 'E11', packing_list['quotation_no'])
safe_write(ws, 'I11', packing_list['quotation_date'])
safe_write(ws, 'I13', packing_list['enquiry_date'])
safe_write(ws, 'I15', packing_list['enquiry_reference'])
ws['A16'] = packing_list['party']
ws['A17'] = address_line1
ws['A17'].alignment = Alignment(horizontal='left')
ws['A17'].alignment = Alignment(wrap_text=True)
ws['A19'] = gst_number

# Packing details section
start_row = 24
num_rows = len(packing_list['packing_details'])
if num_rows > 1:
    ws.insert_rows(start_row + 1, num_rows - 1)

total_crates = total_pcs = total_net = total_gross = total_volume = 0
for i, item in enumerate(packing_list['packing_details']):
    row = start_row + i
    height_display = item.get('crate_height_display', item['crate_height'])
    length_display = item.get('length_display', item['length_of_profiles'] + 100)
    ws[f'A{row}'] = item['description']
    ws[f'B{row}'] = item['no_of_crates']
    ws[f'C{row}'] = length_display 
    ws[f'D{row}'] = item['crate_width']
    ws[f'E{row}'] = height_display
    ws[f'F{row}'] = item['profiles_per_crate']
    ws[f'G{row}'] = round(item['net_weight_per_crate'],0)
    ws[f'H{row}'] = round(item['gross_weight_per_crate'],0)
    ws[f'I{row}'] = item['total_pcs']
    volume_crate = round(
        item['no_of_crates'] * length_display * item['crate_width'] * height_display / 1000000000,
        2
    )
    ws[f'J{row}'] = volume_crate
    ws[f'K{row}'] = item['total_net_weight']
    ws[f'L{row}'] = item['total_gross_weight']
    for col in "ABCDEFGHIJKL":
        ws[f'{col}{row}'].border = thin_border
    total_crates += item['no_of_crates']
    total_pcs += item['total_pcs']
    total_net += item['total_net_weight']
    total_gross += item['total_gross_weight']
    total_volume += volume_crate
    
fill_color = PatternFill(fill_type="solid", fgColor="87CEEB")  # Use any hex color
# Total row
total_row = start_row + num_rows
ws[f'A{total_row}'] = "TOTAL"
ws[f'B{total_row}'] = total_crates
ws[f'I{total_row}'] = total_pcs
ws[f'J{total_row}'] = total_volume
ws[f'K{total_row}'] = total_net
ws[f'L{total_row}'] = total_gross
for col in "ABIJKL":
    ws[f'{col}{total_row}'].font = bold_font
for col in "ABCDEFGHIJKL":
    ws[f'{col}{total_row}'].border = thin_border
    ws[f'{col}{total_row}'].fill = fill_color

# Trailer Loading Details
trailer_start = total_row + 3
ws[f'A{trailer_start}'] = "LOADING DETAILS PER TRAILER BASIS"
ws.merge_cells(f'A{trailer_start}:A{trailer_start + 1}')
ws[f'B{trailer_start}'] = "NO OF TRAILERS"
ws.merge_cells(f'B{trailer_start}:B{trailer_start + 1}')
ws[f'C{trailer_start}'] = "BED WIDTH"
ws.merge_cells(f'C{trailer_start}:D{trailer_start}')
ws[f'E{trailer_start}'] = "HEIGHT"
ws.merge_cells(f'E{trailer_start}:F{trailer_start}')
ws[f'G{trailer_start}'] = "LENGTH"
ws.merge_cells(f'G{trailer_start}:H{trailer_start}')
ws[f'I{trailer_start}'] = "WEIGHT"
ws[f'I{trailer_start + 1}'] = "kg"
for col in "CEG":
   ws[f'{col}{trailer_start + 1}'] = 'mm'
for col in "DFH":
   ws[f'{col}{trailer_start + 1}'] = 'feet'   
fill_color = PatternFill(fill_type="solid", fgColor="A3BDF8")  # Use any hex color   
# for col in "ABCDEFGHI":
for col in range(1,10):   
    col_letter = get_column_letter(col)
    apply_border(ws[f'{col_letter}{trailer_start}'])
    apply_border(ws[f'{col_letter}{trailer_start + 1}'])
    ws[f'{col_letter}{trailer_start}'].fill = fill_color
    ws[f'{col_letter}{trailer_start + 1}'].fill = fill_color
    
trailer_start += 2
for i, (type_name, detail) in enumerate(packing_list['trailer_loading_details'].items()):
    row = trailer_start + i
    ws[f'A{row}'] = type_name
    ws[f'B{row}'] = detail['Used_in_trailers']
    ws[f'C{row}'] = detail['Max_Width_mm'] 
    ws[f'D{row}'] = round(detail['Max_Width_mm'] * 0.0032808399,2)
    ws[f'E{row}'] = detail['Max_Height_mm']
    ws[f'F{row}'] = round(detail['Max_Height_mm'] * 0.0032808399,2)
    ws[f'G{row}'] = detail['Max_Length_mm']
    ws[f'H{row}'] = round(detail['Max_Length_mm'] * 0.0032808399,2)
    ws[f'I{row}'] = detail['Gross_Weight_per_trailer_kg']
    for col in "ABCDEFGHI":
        ws[f'{col}{row}'].border = thin_border
fill_color = PatternFill(fill_type="solid", fgColor="D9D9D9")  # Use any hex color 
row +=2
ws[f'A{row}'] = "TRAILER WISE LOADING ARRANGEMENTS OF CRATES"
ws[f'A{row}'].fill = fill_color
# ws[f'C{row}'] = "Crate arranments per trailer"
ws.merge_cells(f'A{row}:I{row}')
for col in "ABCDEFGHI":
    ws[f'{col}{row}'].border = thin_border
for i, (type_name, detail) in enumerate(packing_list['trailer_loading_details'].items()):
    row = row + 1
    ws[f'A{row}'] = type_name
    ws[f'B{row}'] = detail['Used_in_trailers']
    ws[f'C{row}'] = detail['Crate_Summary']
    ws.merge_cells(f'C{row}:I{row}')
    for col in "ABCDEFGHI":
        ws[f'{col}{row}'].border = thin_border


full_path = os.path.join(folder, new_file)
# Create a workbook and save to the path
wb.save(full_path)
print(f" File saved to: {full_path}")        
# Save file
# wb.save(new_file)
# print(f" Packing list saved to: {new_file}")
for item in main_items:
    total_packing_weight = 0
    for rows in item['crate_details'].values():
        packing_weight = rows['total_angle_weight_kg'] + rows['total_channel_weight_kg']
        total_packing_weight = packing_weight  +  total_packing_weight 
    item['total_packing_wt'] =  total_packing_weight

# - --------------------------------------------------------------------------QUOTATION PRINTING STARTS ----------------------------------------------------------        
# Required inputs (replace with actual values from your program)
template_file = f"{format_folder}/quote_format.xlsx"
new_file1 = f"{packing_list['quotation_no']}.xlsx"
full_path = os.path.join(folder, new_file1)

# Step 1: Copy template (preserves header/footer images)
shutil.copy(template_file, full_path)

# Step 2: Open copied file using Excel automation
app = xw.App(visible=False)
wb = app.books.open(full_path)
ws = wb.sheets[0]

# Step 3: Fill header details
ws.range('E9').value = f"Offer No. : {packing_list['quotation_no']}"
ws.range('E10').value =  f"Offer Dated: {packing_list['quotation_date']}"
ws.range('E11').value = f"Reference: {packing_list['enquiry_reference']}"
ws.range('E12').value = f"Ref dated: {packing_list['enquiry_date']}"

ws.range('A9').value = packing_list['party']
ws.range('A10').value = address_line1
ws.range('A12').value = f"Contact No.: {contact_no}"
ws.range('A13').value = f"Kind Attention : {contact_person}"
ws.range('A13').api.Font.Bold = True
edges = [
    constants.xlEdgeLeft,
    constants.xlEdgeTop,
    constants.xlEdgeBottom,
    constants.xlEdgeRight ]
row = 20
num_rows = len(main_items)
for i, item in enumerate(main_items):
   
   ws.range(f'A{row}').value = i + 1
   spec = input(f"Standard Specification of Item No. {i+1}? (y /n ):  ")
   if spec =="y" or spec == "":
       specs = item['bom'][0]['specs']
   else: specs = input("Enter Specification : ")   
   ws.range(f'B{row}').value = f"{item['desc']}:                                   MOC {specs}"
   ws.range(f'C{row}').value = f"{round(item['width_OD'],2)} X {round(item['thks'],2)} x {int(round(item['t_length'],0))}"
   ws.range(f'D{row}').value = item['qty']
   
   ws.range(f'E{row}').value = f"{int(round(item['rate_offered_per_pc'],0))}/-"
   ws.range(f'E{row}').api.Font.Bold = True
   ws.range(f'F{row}').value = f"{int(round(item['qty'] * item['rate_offered_per_pc'],0))}/-"
   ws.range(f'F{row}').api.Font.Bold = True
   
   words = "Rupees " + num2words(item['rate_offered_per_pc'], lang='en_IN') + " only"
   words = words.title()
   ws.range(f'A{row}').api.HorizontalAlignment = -4108  # xlCenter
   ws.range(f'A{row}').api.VerticalAlignment = -4108    # xlCenter
   ws.range(f'B{row}').api.HorizontalAlignment = -4108
   ws.range(f'B{row}').api.VerticalAlignment = -4108
   for j in range(2):
      for col in "ABCDEF":
         cell1 = ws.range(f"{col}{row+j-1}")
         for edge in edges:
            border = cell1.api.Borders(edge)
            border.LineStyle = constants.xlContinuous
            border.Weight = constants.xlThin
            border.ColorIndex = 0  # black
   row +=1
# Step 7: Save and close
wb.save()
wb.close()
app.quit()

print(f"Quotation saved with header/footer to: {full_path}")





   
            
