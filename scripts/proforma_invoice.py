import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook


def write_header(ws, title):
    ws["A1"] = title
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def write_kv(ws, start_row, data):
    row = start_row
    for key, value in data:
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1
    return row


def main():
    try:
        payload = json.load(sys.stdin)
    except json.JSONDecodeError:
        print(json.dumps({"ok": False, "error": "Invalid JSON payload"}))
        return 1

    output_path = payload.get("outputPath")
    if not output_path:
        print(json.dumps({"ok": False, "error": "Missing outputPath"}))
        return 1

    order = payload.get("order") or {}
    quotation = payload.get("quotation") or {}
    inquiries = payload.get("inquiries") or []

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"
    write_header(ws, "Proforma Invoice")

    row = 4
    row = write_kv(
        ws,
        row,
        [
            ("Offer No", order.get("offerNo", "")),
            ("Offer Date", order.get("offerDate", "")),
            ("PO Number", order.get("poNumber", "")),
            ("PO Date", order.get("poDate", "")),
            ("Party Name", order.get("partyName", "")),
            ("Party Address", quotation.get("partyAddress", "")),
            ("Contact No", quotation.get("contactNo", "")),
        ],
    )

    if inquiries:
        row += 2
        ws[f"A{row}"] = "Invoice Items"
        row += 1
        ws[f"A{row}"] = "Item"
        ws[f"B{row}"] = "Dimensions"
        ws[f"C{row}"] = "Quantity"
        ws[f"D{row}"] = "Rate/Pc"
        row += 1
        for inquiry in inquiries:
            dimensions = f"{inquiry.get('widthOd', '')} X {inquiry.get('thickness', '')} X {inquiry.get('length', '')}"
            ws[f"A{row}"] = inquiry.get("item", "")
            ws[f"B{row}"] = dimensions.strip()
            ws[f"C{row}"] = inquiry.get("quantity", "")
            ws[f"D{row}"] = inquiry.get("ratePerPiece", "")
            row += 1

    wb.save(output_path)
    print(json.dumps({"ok": True, "filePath": output_path}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
